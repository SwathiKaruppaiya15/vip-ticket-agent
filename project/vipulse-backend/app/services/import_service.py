"""
Import Service — Excel and PDF ticket import with AI extraction.

Excel:  pandas + openpyxl parse .xlsx/.xls/.csv, validate schema, bulk insert.
PDF:    pdfplumber extracts text, Groq LLM parses structured ticket data.
"""
import asyncio
import io
import re
import json
from typing import Any, Optional
from uuid import uuid4

import structlog

from app.core.config import settings
from app.models.ticket import Ticket
from app.schemas.ticket_schemas import TicketCreateRequest
from app.services.ticket_service import TicketService

logger = structlog.get_logger(__name__)
_svc = TicketService()

# ── Constants ─────────────────────────────────────────────────────────────────

REQUIRED_COLUMNS = {
    "employee_id", "employee_name", "role",
    "department", "issue_title", "issue_description", "severity",
}

VALID_SEVERITIES = {"low", "medium", "high", "critical"}

# Column alias mapping (case-insensitive)
COLUMN_ALIASES: dict[str, str] = {
    # employee_id aliases
    "emp_id": "employee_id", "empid": "employee_id", "id": "employee_id",
    "employee id": "employee_id",
    # employee_name aliases
    "emp_name": "employee_name", "name": "employee_name", "full_name": "employee_name",
    "employee name": "employee_name",
    # role aliases
    "job_title": "role", "title": "role", "designation": "role",
    # department aliases
    "dept": "department", "team": "department", "division": "department",
    # issue_title aliases
    "subject": "issue_title", "title": "issue_title", "summary": "issue_title",
    "issue title": "issue_title",
    # issue_description aliases
    "description": "issue_description", "desc": "issue_description",
    "details": "issue_description", "issue_details": "issue_description",
    "issue description": "issue_description",
    # severity aliases
    "priority": "severity", "urgency": "severity", "level": "severity",
}

# Severity normalisation — maps common variants to valid values
SEVERITY_MAP: dict[str, str] = {
    "low":      "low",  "l":  "low",  "1": "low",
    "medium":   "medium","med": "medium","m": "medium","moderate":"medium","2":"medium",
    "high":     "high", "h":  "high", "3": "high",
    "critical": "critical","crit":"critical","c":"critical","urgent":"critical","4":"critical",
}


# ─────────────────────────────────────────────────────────────────────────────
# Excel / CSV Import
# ─────────────────────────────────────────────────────────────────────────────

class ExcelImportResult:
    __slots__ = ("total_rows", "created", "failed", "errors", "ticket_ids")

    def __init__(self) -> None:
        self.total_rows: int = 0
        self.created:    int = 0
        self.failed:     int = 0
        self.errors:     list[dict[str, Any]] = []
        self.ticket_ids: list[str] = []

    def to_dict(self) -> dict:
        return {
            "total_rows": self.total_rows,
            "created":    self.created,
            "failed":     self.failed,
            "errors":     self.errors,
            "ticket_ids": self.ticket_ids,
        }


def _normalise_columns(df_columns: list[str]) -> dict[str, str]:
    """
    Build a mapping: original_column_name → canonical_field_name.
    Strips whitespace, lowercases, resolves aliases.
    """
    mapping: dict[str, str] = {}
    for col in df_columns:
        clean = col.strip().lower().replace("-", "_").replace(" ", "_")
        # Direct match
        if clean in REQUIRED_COLUMNS:
            mapping[col] = clean
            continue
        # Alias match (original cleaned, including spaces)
        clean_spaced = col.strip().lower()
        if clean_spaced in COLUMN_ALIASES:
            mapping[col] = COLUMN_ALIASES[clean_spaced]
        elif clean in COLUMN_ALIASES:
            mapping[col] = COLUMN_ALIASES[clean]
    return mapping


def _normalise_severity(raw: str) -> Optional[str]:
    if not raw:
        return None
    return SEVERITY_MAP.get(str(raw).strip().lower())


def parse_excel_bytes(
    file_bytes: bytes,
    filename: str,
    *,
    max_rows: int = 1000,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """
    Parse bytes from .xlsx/.xls/.csv into (valid_rows, error_rows).

    valid_rows   = list of dicts ready for TicketCreateRequest
    error_rows   = list of {"row": N, "error": "..."}
    """
    try:
        import pandas as pd
    except ImportError:
        raise RuntimeError("pandas is required for Excel import. Run: pip install pandas openpyxl")

    # ── Parse ──────────────────────────────────────────────────────────────────
    buf = io.BytesIO(file_bytes)
    ext = filename.rsplit(".", 1)[-1].lower()

    try:
        if ext == "csv":
            df = pd.read_csv(buf, dtype=str, keep_default_na=False)
        elif ext in ("xlsx", "xls"):
            df = pd.read_excel(buf, dtype=str, keep_default_na=False)
        else:
            raise ValueError(f"Unsupported file type: {ext}")
    except Exception as exc:
        raise ValueError(f"Cannot parse file: {exc}") from exc

    if len(df) > max_rows:
        df = df.iloc[:max_rows]

    # ── Map columns ────────────────────────────────────────────────────────────
    col_map = _normalise_columns(list(df.columns))
    df = df.rename(columns=col_map)

    missing = REQUIRED_COLUMNS - set(df.columns)
    if missing:
        raise ValueError(
            f"Missing required columns: {', '.join(sorted(missing))}. "
            f"Got: {', '.join(df.columns)}"
        )

    valid_rows: list[dict[str, Any]] = []
    error_rows: list[dict[str, Any]] = []

    for idx, row in df.iterrows():
        row_num = int(idx) + 2  # 1-indexed, +1 for header

        # Skip completely empty rows
        values = [str(row.get(c, "")).strip() for c in REQUIRED_COLUMNS]
        if not any(values):
            continue

        errors: list[str] = []

        # Validate required fields
        for field in REQUIRED_COLUMNS:
            val = str(row.get(field, "")).strip()
            if not val:
                errors.append(f"'{field}' is required")

        # Validate severity
        raw_sev = str(row.get("severity", "")).strip()
        norm_sev = _normalise_severity(raw_sev)
        if norm_sev is None:
            errors.append(
                f"Invalid severity '{raw_sev}'. Allowed: low, medium, high, critical"
            )

        # Validate lengths
        issue_title = str(row.get("issue_title", "")).strip()
        if issue_title and len(issue_title) < 5:
            errors.append("issue_title must be at least 5 characters")

        issue_desc = str(row.get("issue_description", "")).strip()
        if issue_desc and len(issue_desc) < 10:
            errors.append("issue_description must be at least 10 characters")

        if errors:
            error_rows.append({"row": row_num, "error": "; ".join(errors)})
            continue

        valid_rows.append({
            "employee_id":       str(row.get("employee_id", "")).strip(),
            "employee_name":     str(row.get("employee_name", "")).strip(),
            "role":              str(row.get("role", "")).strip(),
            "department":        str(row.get("department", "")).strip(),
            "issue_title":       issue_title,
            "issue_description": issue_desc,
            "severity":          norm_sev,
        })

    return valid_rows, error_rows


async def import_excel_tickets(
    file_bytes: bytes,
    filename: str,
    created_by: str,
) -> ExcelImportResult:
    """
    Full Excel import pipeline:
      1. Parse & validate rows
      2. Bulk insert valid tickets (each gets background AI pipeline)
      3. Return result summary
    """
    result = ExcelImportResult()

    try:
        valid_rows, error_rows = parse_excel_bytes(file_bytes, filename)
    except ValueError as exc:
        result.total_rows = 0
        result.failed = 1
        result.errors = [{"row": 0, "error": str(exc)}]
        return result

    result.total_rows = len(valid_rows) + len(error_rows)
    result.errors = error_rows
    result.failed = len(error_rows)

    if not valid_rows:
        return result

    # Insert tickets concurrently (capped at 20 concurrent to avoid DB overload)
    sem = asyncio.Semaphore(20)

    async def _insert_one(row_data: dict) -> Optional[str]:
        async with sem:
            try:
                req = TicketCreateRequest(**row_data)
                ticket = await _svc.create_ticket(req, created_by=created_by)
                # Fire AI pipeline asynchronously — don't wait
                asyncio.create_task(_run_pipeline(ticket))
                return ticket.ticket_id
            except Exception as exc:
                logger.warning("excel_row_insert_failed", error=str(exc), row=row_data)
                return None

    tasks = [_insert_one(row) for row in valid_rows]
    ticket_ids = await asyncio.gather(*tasks)

    for tid in ticket_ids:
        if tid:
            result.created += 1
            result.ticket_ids.append(tid)
        else:
            result.failed += 1
            result.errors.append({"row": -1, "error": "Database insert failed"})

    return result


async def _run_pipeline(ticket: Ticket) -> None:
    """Fire-and-forget AI pipeline for an imported ticket."""
    try:
        await _svc.run_pipeline_and_update(ticket)
    except Exception as exc:
        logger.warning("import_pipeline_error", ticket_id=ticket.ticket_id, error=str(exc))


# ─────────────────────────────────────────────────────────────────────────────
# PDF / Image Import
# ─────────────────────────────────────────────────────────────────────────────

EXTRACT_PROMPT = """
You are an IT helpdesk AI assistant. Extract ticket information from the following text.

Return ONLY a valid JSON array of ticket objects. Each object must have exactly these fields:
- employee_name: string (required, use "Unknown" if not found)
- role: string (required, use "Employee" if not found)  
- department: string (required, use "General" if not found)
- issue_title: string (required, min 5 chars, concise summary)
- issue_description: string (required, min 20 chars, detailed description)
- severity: string (must be one of: low, medium, high, critical)

Severity inference rules:
- CEO, CFO, CTO, Director, VP, C-suite roles → increase severity by one level
- Keywords: "server down", "production outage", "security breach", "payment failure",
  "VPN failure", "data loss", "critical system", "cannot login", "breach" → high or critical
- "urgent", "ASAP", "emergency" → high
- Performance issues, slow systems → medium
- General queries, minor issues → low

If the text contains multiple distinct issues, return one object per issue (max 20).
If text is unclear or no ticket-worthy content, return an empty array [].

Text to analyze:
{text}

Return ONLY the JSON array, no other text, no markdown code blocks.
""".strip()


def _extract_text_from_pdf(file_bytes: bytes) -> str:
    """Extract text from PDF using pdfplumber with pytesseract OCR fallback."""
    text = ""

    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            pages_text = []
            for page in pdf.pages[:20]:  # Limit to 20 pages
                page_text = page.extract_text() or ""
                pages_text.append(page_text)
            text = "\n\n".join(pages_text).strip()
    except ImportError:
        logger.warning("pdfplumber not installed, trying PyMuPDF")
    except Exception as exc:
        logger.warning("pdfplumber_error", error=str(exc))

    # Try PyMuPDF if pdfplumber gave nothing
    if not text:
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(stream=file_bytes, filetype="pdf")
            pages_text = [doc[i].get_text() for i in range(min(len(doc), 20))]
            text = "\n\n".join(pages_text).strip()
            doc.close()
        except ImportError:
            logger.warning("PyMuPDF not installed")
        except Exception as exc:
            logger.warning("pymupdf_error", error=str(exc))

    return text


def _extract_text_from_image(file_bytes: bytes, mime_type: str) -> str:
    """Extract text from image using pytesseract OCR."""
    try:
        import pytesseract
        from PIL import Image
        img = Image.open(io.BytesIO(file_bytes))
        return pytesseract.image_to_string(img).strip()
    except ImportError:
        logger.warning("pytesseract/PIL not installed — cannot OCR image")
        return ""
    except Exception as exc:
        logger.warning("ocr_error", error=str(exc))
        return ""


def _call_groq_for_extraction(text: str) -> list[dict[str, Any]]:
    """Call Groq LLM to extract structured ticket data from text."""
    if not settings.GROQ_API_KEY:
        raise RuntimeError("GROQ_API_KEY not configured")

    try:
        from langchain_groq import ChatGroq
        from langchain_core.messages import HumanMessage
    except ImportError:
        raise RuntimeError("langchain_groq not installed")

    # Truncate text to avoid token limits (~8k chars ≈ ~2k tokens)
    truncated = text[:8000] if len(text) > 8000 else text

    llm = ChatGroq(
        model=settings.GROQ_MODEL,
        api_key=settings.GROQ_API_KEY,
        temperature=0.1,
        max_tokens=2048,
    )

    prompt = EXTRACT_PROMPT.format(text=truncated)
    response = llm.invoke([HumanMessage(content=prompt)])
    raw = response.content.strip()

    # Strip markdown code blocks if LLM wraps response
    raw = re.sub(r"^```(?:json)?\s*", "", raw, flags=re.IGNORECASE)
    raw = re.sub(r"\s*```$", "", raw)
    raw = raw.strip()

    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError as exc:
        logger.warning("llm_json_parse_error", raw=raw[:200], error=str(exc))
        # Try to extract JSON array from response
        match = re.search(r"\[.*\]", raw, re.DOTALL)
        if match:
            parsed = json.loads(match.group(0))
        else:
            return []

    if not isinstance(parsed, list):
        parsed = [parsed] if isinstance(parsed, dict) else []

    return parsed


def _validate_extracted_tickets(
    raw_list: list[dict[str, Any]]
) -> tuple[list[dict[str, Any]], list[str]]:
    """Validate and normalise extracted ticket data. Returns (valid, errors)."""
    valid: list[dict[str, Any]] = []
    errors: list[str] = []

    for i, item in enumerate(raw_list[:20]):
        if not isinstance(item, dict):
            errors.append(f"Item {i+1}: not a valid object")
            continue

        # Normalise severity
        sev = _normalise_severity(str(item.get("severity", "medium")))
        if sev is None:
            sev = "medium"

        # Build validated item
        ticket = {
            "employee_name":     str(item.get("employee_name", "Unknown")).strip() or "Unknown",
            "role":              str(item.get("role", "Employee")).strip() or "Employee",
            "department":        str(item.get("department", "General")).strip() or "General",
            "issue_title":       str(item.get("issue_title", "")).strip(),
            "issue_description": str(item.get("issue_description", "")).strip(),
            "severity":          sev,
        }

        if len(ticket["issue_title"]) < 5:
            errors.append(f"Item {i+1}: issue_title too short")
            continue

        if len(ticket["issue_description"]) < 10:
            # Try to use title as description if description is empty
            if not ticket["issue_description"]:
                ticket["issue_description"] = ticket["issue_title"] + " — requires immediate attention."
            else:
                errors.append(f"Item {i+1}: issue_description too short")
                continue

        valid.append(ticket)

    return valid, errors


async def parse_pdf_or_image(
    file_bytes: bytes,
    filename: str,
    mime_type: str,
) -> tuple[list[dict[str, Any]], str]:
    """
    Extract ticket data from PDF or image using AI.

    Returns:
        (validated_tickets, extracted_text)
    validated_tickets: ready-to-preview ticket dicts
    extracted_text:    raw text (for debugging / display)
    """
    # 1. Extract text
    if mime_type == "application/pdf" or filename.lower().endswith(".pdf"):
        text = _extract_text_from_pdf(file_bytes)
    else:
        text = _extract_text_from_image(file_bytes, mime_type)

    if not text or len(text.strip()) < 10:
        return [], text

    # 2. Call LLM — run in thread executor so we don't block the event loop
    loop = asyncio.get_running_loop()
    try:
        raw_tickets = await loop.run_in_executor(None, _call_groq_for_extraction, text)
    except Exception as exc:
        logger.error("groq_extraction_failed", error=str(exc))
        raise RuntimeError(f"AI extraction failed: {exc}") from exc

    # 3. Validate
    validated, _ = _validate_extracted_tickets(raw_tickets)
    return validated, text


async def create_pdf_tickets(
    extracted_tickets: list[dict[str, Any]],
    employee_id: str,
    created_by: str,
) -> ExcelImportResult:
    """
    Bulk-create tickets from PDF extraction results.
    employee_id is used since PDF may not contain it.
    """
    result = ExcelImportResult()
    result.total_rows = len(extracted_tickets)

    sem = asyncio.Semaphore(10)

    async def _insert(item: dict) -> Optional[str]:
        async with sem:
            try:
                req = TicketCreateRequest(
                    employee_id=employee_id or f"PDF-{uuid4().hex[:6].upper()}",
                    **item,
                )
                ticket = await _svc.create_ticket(req, created_by=created_by)
                asyncio.create_task(_run_pipeline(ticket))
                return ticket.ticket_id
            except Exception as exc:
                logger.warning("pdf_row_insert_failed", error=str(exc))
                return None

    ids = await asyncio.gather(*[_insert(t) for t in extracted_tickets])
    for tid in ids:
        if tid:
            result.created += 1
            result.ticket_ids.append(tid)
        else:
            result.failed += 1
            result.errors.append({"row": -1, "error": "Insert failed"})

    return result


# ─────────────────────────────────────────────────────────────────────────────
# Sample Template Generator
# ─────────────────────────────────────────────────────────────────────────────

def generate_sample_excel() -> bytes:
    """Generate a sample .xlsx template with headers and 3 example rows."""
    try:
        import pandas as pd
    except ImportError:
        raise RuntimeError("pandas is required. Run: pip install pandas openpyxl")

    sample_data = [
        {
            "employee_id":       "EMP-001",
            "employee_name":     "Alice Johnson",
            "role":              "Senior Engineer",
            "department":        "Engineering",
            "issue_title":       "VPN connection keeps dropping",
            "issue_description": "My VPN disconnects every 20-30 minutes making it impossible to work remotely. Already tried reinstalling the client twice.",
            "severity":          "high",
        },
        {
            "employee_id":       "EMP-042",
            "employee_name":     "Robert Chen",
            "role":              "Finance Manager",
            "department":        "Finance",
            "issue_title":       "Cannot access payroll system",
            "issue_description": "Getting 403 Forbidden error when trying to access the payroll dashboard since this morning. Month-end processing is blocked.",
            "severity":          "critical",
        },
        {
            "employee_id":       "EMP-101",
            "employee_name":     "Sarah Williams",
            "role":              "Marketing Analyst",
            "department":        "Marketing",
            "issue_title":       "Outlook keeps crashing on startup",
            "issue_description": "Microsoft Outlook crashes immediately after loading. Have tried safe mode and disabling add-ins but the issue persists.",
            "severity":          "medium",
        },
    ]

    df = pd.DataFrame(sample_data)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Tickets")

        # Style the header row
        try:
            ws = writer.sheets["Tickets"]
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill   = header_fill
                cell.font   = header_font
                cell.alignment = Alignment(horizontal="center")
            # Auto-width columns
            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
        except Exception:
            pass  # Styling is optional

    return buf.getvalue()
